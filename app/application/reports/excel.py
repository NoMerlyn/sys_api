"""Excel (XLSX) export for sales reports.

Uses openpyxl to produce a workbook with one sheet per report
(summary, top products, top clients, top sellers, by day). The
workbook is returned as bytes ready to be served as
`application/vnd.openxmlformats-officedocument.spreadsheetml.sheet`.
"""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Light formatting so the file is readable when opened in Excel / LibreOffice.
_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_BOLD_FONT = Font(bold=True)
_CENTER = Alignment(horizontal="center")


def _autosize(ws) -> None:
    """Resize each column to fit its longest cell."""
    for col_cells in ws.columns:
        col = col_cells[0].column
        if col is None:
            continue
        letter = get_column_letter(col)
        max_len = 0
        for cell in col_cells:
            v = cell.value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = min(max_len + 2, 40)


def _write_header(ws, headers: list[str]) -> None:
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = _CENTER


def build_summary_sheet(wb: Workbook, summary: dict[str, Any]) -> None:
    ws = wb.active
    ws.title = "Resumen"
    rows = [
        ("Total de facturas (confirmadas)", int(summary.get("total_invoices", 0))),
        ("Ingresos totales", float(summary.get("total_amount", 0))),
        ("Promedio por factura", float(summary.get("avg_amount", 0))),
        ("Clientes únicos", int(summary.get("distinct_clients", 0))),
        ("Vendedores", int(summary.get("distinct_sellers", 0))),
        ("", ""),
        ("Por estado", ""),
    ]
    by_status = summary.get("by_status") or {}
    for status, count in by_status.items():
        rows.append((f"  {status}", int(count)))
    for i, (label, value) in enumerate(rows, start=1):
        c_label = ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=value)
        if label and not value and not label.startswith("  "):
            c_label.font = _BOLD_FONT
    _autosize(ws)


def _rows_sheet(
    wb: Workbook,
    *,
    title: str,
    headers: list[str],
    rows: list[list[Any]],
) -> None:
    ws = wb.create_sheet(title)
    _write_header(ws, headers)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    _autosize(ws)


def build_top_products_sheet(wb: Workbook, items: list[dict[str, Any]]) -> None:
    rows = [
        [
            i + 1,
            item.get("product_name") or "—",
            float(item.get("quantity", 0)),
            int(item.get("invoices", 0)),
            float(item.get("revenue", 0)),
        ]
        for i, item in enumerate(items)
    ]
    _rows_sheet(
        wb,
        title="Top productos",
        headers=["#", "Producto", "Cantidad", "Facturas", "Ingresos"],
        rows=rows,
    )


def build_top_clients_sheet(wb: Workbook, items: list[dict[str, Any]]) -> None:
    rows = [
        [
            i + 1,
            item.get("client_name") or "—",
            int(item.get("invoices", 0)),
            float(item.get("spent", 0)),
        ]
        for i, item in enumerate(items)
    ]
    _rows_sheet(
        wb,
        title="Top clientes",
        headers=["#", "Cliente", "Facturas", "Total gastado"],
        rows=rows,
    )


def build_top_sellers_sheet(wb: Workbook, items: list[dict[str, Any]]) -> None:
    rows = [
        [
            i + 1,
            f'{item.get("name", "")} {item.get("last_name", "")}'.strip() or "—",
            item.get("username", ""),
            int(item.get("invoices", 0)),
            float(item.get("sold", 0)),
        ]
        for i, item in enumerate(items)
    ]
    _rows_sheet(
        wb,
        title="Top vendedores",
        headers=["#", "Vendedor", "Username", "Facturas", "Total vendido"],
        rows=rows,
    )


def build_by_day_sheet(wb: Workbook, items: list[dict[str, Any]]) -> None:
    rows = [
        [item.get("day", ""), int(item.get("invoices", 0)), float(item.get("revenue", 0))]
        for item in items
    ]
    _rows_sheet(
        wb,
        title="Ingresos por día",
        headers=["Día", "Facturas", "Ingresos"],
        rows=rows,
    )


def build_workbook(
    *,
    summary: dict[str, Any],
    top_products: list[dict[str, Any]],
    top_clients: list[dict[str, Any]],
    top_sellers: list[dict[str, Any]],
    by_day: list[dict[str, Any]],
) -> bytes:
    """Build the full XLSX as bytes."""
    wb = Workbook()
    build_summary_sheet(wb, summary)
    build_top_products_sheet(wb, top_products)
    build_top_clients_sheet(wb, top_clients)
    build_top_sellers_sheet(wb, top_sellers)
    build_by_day_sheet(wb, by_day)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
