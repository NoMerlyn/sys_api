"""Excel export builder tests (no DB)."""

from __future__ import annotations

from openpyxl import load_workbook

from app.application.reports.excel import build_workbook


def test_build_workbook_returns_a_valid_xlsx() -> None:
    blob = build_workbook(
        summary={
            "total_invoices": 3,
            "total_amount": 336.0,
            "avg_amount": 112.0,
            "distinct_clients": 1,
            "distinct_sellers": 1,
            "by_status": {"CONFIRMED": 3, "DRAFT": 1},
        },
        top_products=[
            {"product_id": 1, "product_name": "Pen", "quantity": 4, "revenue": 200.0, "invoices": 2},
            {"product_id": 2, "product_name": "Book", "quantity": 2, "revenue": 100.0, "invoices": 1},
        ],
        top_clients=[
            {"client_id": 1, "client_name": "Consumidor final", "invoices": 3, "spent": 336.0},
        ],
        top_sellers=[
            {"user_id": 1, "username": "u1", "name": "S", "last_name": "O", "invoices": 3, "sold": 336.0},
        ],
        by_day=[
            {"day": "2026-01-01", "invoices": 1, "revenue": 112.0},
            {"day": "2026-01-02", "invoices": 1, "revenue": 112.0},
        ],
    )
    # XLSX is a zip; check the magic bytes.
    assert blob[:2] == b"PK"
    # The blob should load with openpyxl.
    wb = load_workbook(filename=__import__("io").BytesIO(blob))
    assert set(wb.sheetnames) == {"Resumen", "Top productos", "Top clientes", "Top vendedores", "Ingresos por día"}
    # Top productos: row 2 = Pen, row 3 = Book
    top_products = wb["Top productos"]
    assert top_products.cell(row=2, column=1).value == 1
    assert top_products.cell(row=2, column=2).value == "Pen"
    assert top_products.cell(row=2, column=5).value == 200.0
    # Resumen has 7 rows (5 metrics + blank + header for by_status)
    assert wb["Resumen"].max_row >= 7
